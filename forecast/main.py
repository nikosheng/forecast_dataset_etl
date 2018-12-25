# -*- coding: utf-8 -*-

__author__ = "Niko Feng"
__license__ = "GPL"
__version__ = "1.0.0"

#######################################################
# Forecast Dataset ETL Script
#######################################################

from openpyxl import load_workbook, Workbook
from functools import reduce
from collections import Counter
from operator import itemgetter
import getopt
import sys
import os


class Prod:
    def __init__(self, product_name, product_code, trade_time, category,
                 sell_type, price, is_fresh, is_weigh, discount_price, is_discount):
        self.product_name = product_name
        self.product_code = product_code
        self.trade_time = trade_time
        self.category = category
        self.sell_type = sell_type
        self.price = price
        self.is_fresh = is_fresh
        self.is_weigh = is_weigh
        self.discount_price = discount_price
        self.is_discount = is_discount


class ConvertedProd:
    def __init__(self, product_name, product_code, trade_period, sum, category,
                 sell_type, price, is_fresh, is_weigh, discount_price, is_discount, _i=0):
        self._i = _i
        self.product_name = product_name
        self.product_code = product_code
        self.trade_period = trade_period
        self.sum = sum
        self.category = category
        self.sell_type = sell_type
        self.price = price
        self.is_fresh = is_fresh
        self.is_weigh = is_weigh
        self.discount_price = discount_price
        self.is_discount = is_discount

    def __iter__(self):
        return self

    def __next__(self):
        if self._i == 0:
            self._i += 1
            return self.product_name
        elif self._i == 1:
            self._i += 1
            return self.product_code
        elif self._i == 2:
            self._i += 1
            return self.trade_period
        elif self._i == 3:
            self._i += 1
            return self.sum
        elif self._i == 4:
            self._i += 1
            return self.category
        elif self._i == 5:
            self._i += 1
            return self.sell_type
        elif self._i == 6:
            self._i += 1
            return self.price
        elif self._i == 7:
            self._i += 1
            return self.is_fresh
        elif self._i == 8:
            self._i += 1
            return self.is_weigh
        elif self._i == 9:
            self._i += 1
            return self.discount_price
        elif self._i == 10:
            self._i += 1
            return self.is_discount
        else:
            raise StopIteration()


def to_product(list):
    """
    Convert to class Prod
    :param list:
    :return:
    """
    product = Prod(product_name=list[0],
                   product_code=list[1],
                   trade_time=list[2],
                   category=list[3],
                   sell_type=list[4],
                   price=list[5],
                   is_fresh=list[6],
                   is_weigh=list[7],
                   discount_price=list[8],
                   is_discount=list[9])
    return product


def to_products(list):
    """
    Convert to class Prod with a list of objects
    :param list:
    :return:
    """
    products = [to_product(product) for product in list]
    return products


def sort_by_trade_time(products):
    """
    Sorted by trade time
    :param products:
    :return:
    """
    sorted_products = sorted(products, key=itemgetter(2), reverse=False)
    return sorted_products


def gen_converted_prod(products):
    """
    Generate the converted product with attributes
    :param products: The product list
    :return:
    """
    # Calculate the average price
    price_list = [prod.price for prod in products]
    average_price = round(reduce(lambda x, y: x + y, price_list) / len(price_list), 2)
    # Choose the discount sell_type based on the max number
    sell_type_list = [prod.sell_type for prod in products]
    sell_type = get_most_common_item(sell_type_list)
    # Choose the discount tag based on the max number
    discount_tag_list = [prod.is_discount for prod in products]
    discount_tag = get_most_common_item(discount_tag_list)

    converted_prod = products[0]
    converted_prod.price = average_price
    converted_prod.sell_type = sell_type
    converted_prod.is_discount = discount_tag
    converted_prod.sum = len(products)
    return converted_prod


def get_most_common_item(list):
    """
    Get the Top 1 item in the list
    :param list:
    :return:
    """
    counter = Counter(list)
    return counter.most_common(1)[0][0]


def export_to_excel_workbook(file_name, products, output):
    """
    Export to excel workbook
    :param file_name: The file name
    :param products: The product list
    :param output: The output folder
    :return:
    """
    wb = Workbook()

    ws = wb.active  # 创建一个sheet
    ws.title = file_name

    ws["A1"] = "商品名称"
    ws["B1"] = "商品编码"
    ws["C1"] = "交易时间"
    ws["D1"] = "总数"
    ws["E1"] = "品类"
    ws["F1"] = "销售类型"
    ws["G1"] = "销售价格"
    ws["H1"] = "是否生鲜"
    ws["I1"] = "是否称重"
    ws["J1"] = "优惠金额"
    ws["K1"] = "会否优惠"

    row_start_num = 2
    column_start_num = 1
    for product in products:
        for element in product:
            ws.cell(row=row_start_num, column=column_start_num, value=element)
            column_start_num += 1
        row_start_num += 1
        column_start_num = 1

    output_file = "{output}/{filename}.xlsx".format(output=output, filename=file_name)
    wb.save(output_file)


def generate_by_hour(list, output):
    """
    Generate the excel workbook by hourly calculation
    :param list:
    :param output:
    :return:
    """
    products = to_products(list)

    # ConvertedProduct List
    converted_products = []

    prev_prod = products[0]
    products_hour_period = []

    # File Name
    file_name = "{product_code}_h".format(product_code=prev_prod.product_code)

    for index in range(len(products)):
        product = products[index]
        converted_prod = ConvertedProd(product_name=product.product_name,
                                       product_code=product.product_code,
                                       category=product.category,
                                       sell_type=product.sell_type,
                                       price=product.price,
                                       is_fresh=product.is_fresh,
                                       is_weigh=product.is_weigh,
                                       discount_price=product.discount_price,
                                       is_discount=product.is_discount,
                                       trade_period="{year}/{month}/{day} {from_hour}-{to_hour}".format(
                                           year=product.trade_time.year,
                                           month=product.trade_time.month,
                                           day=product.trade_time.day,
                                           from_hour=product.trade_time.hour,
                                           to_hour=product.trade_time.hour + 1),
                                       sum=0)
        if product.trade_time.hour == prev_prod.trade_time.hour:
            products_hour_period.append(converted_prod)
            prev_prod = product
        else:
            converted_products.append(gen_converted_prod(products=products_hour_period))
            prev_prod = product
            products_hour_period.clear()
            products_hour_period.append(converted_prod)

    # handle the last product list
    converted_products.append(gen_converted_prod(products=products_hour_period))
    # export to a new workbook based on hourly calculation
    output_dir = output + "/hour"
    if not os.path.exists(output_dir):
        os.mkdir(output_dir)
    export_to_excel_workbook(file_name, converted_products, output_dir)


def generate_by_day(list, output):
    """
    Generate the excel workbook by daily calculation
    :param list:
    :param output:
    :return:
    """
    products = to_products(list)

    # ConvertedProduct List
    converted_products = []

    prev_prod = products[0]
    products_daily_period = []

    # File Name
    file_name = "{product_code}_d".format(product_code=prev_prod.product_code)

    for index in range(len(products)):
        product = products[index]
        converted_prod = ConvertedProd(product_name=product.product_name,
                                       product_code=product.product_code,
                                       category=product.category,
                                       sell_type=product.sell_type,
                                       price=product.price,
                                       is_fresh=product.is_fresh,
                                       is_weigh=product.is_weigh,
                                       discount_price=product.discount_price,
                                       is_discount=product.is_discount,
                                       trade_period="{year}/{month}/{day}".format(
                                           year=product.trade_time.year,
                                           month=product.trade_time.month,
                                           day=product.trade_time.day),
                                       sum=0)
        if product.trade_time.day == prev_prod.trade_time.day:
            products_daily_period.append(converted_prod)
            prev_prod = product
        else:
            converted_products.append(gen_converted_prod(products=products_daily_period))
            prev_prod = product
            products_daily_period.clear()
            products_daily_period.append(converted_prod)

    # handle the last product list
    converted_products.append(gen_converted_prod(products=products_daily_period))
    # export to a new workbook based on hourly calculation
    output_dir = output + "/day"
    if not os.path.exists(output_dir):
        os.mkdir(output_dir)
    export_to_excel_workbook(file_name, converted_products, output_dir)

def help():
    """
    Helper function
    :return:
    """
    print('''
        main.py --import-file <filename> --output-folder <folder>
        -i --import-file  import file name
        -o --output-folder
        ''')


def main(argv):
    """
    Main method
    :param argv: Arguments
    :return:
    """
    try:
        opts, args = getopt.getopt(argv, "hi:o:", ["help", "import-file=", "output-folder"])
    except getopt.GetoptError:
        help()
        sys.exit(2)

    import_file = None
    output_folder = None
    for opt, arg in opts:
        if opt in ("-h", "--help"):
            help()
            sys.exit()
        else:
            if opt in ("-i", "--import-file"):
                import_file = arg
            if opt in ("-o", "--output-folder"):
                output_folder = arg
                if not os.path.exists(output_folder):
                    os.mkdir(output_folder)

    if not os.path.exists(import_file) or not os.path.exists(output_folder):
        sys.exit(2)

    # import workbook
    wb = load_workbook(filename=import_file)
    ws = wb.active
    # Iterate the products and filter the same products
    # Starting from the second row to skip the description
    prev_prod = [cell.value for row in ws.iter_rows(min_row=2, max_row=2) for cell in row]
    if len(prev_prod) == 0:
        sys.exit(2)
    prod_in_same_category = []

    total_row_num = ws.max_row + 1
    for row_num in range(2, total_row_num):
        prod = [cell.value for row in ws.iter_rows(min_row=row_num, max_row=row_num) for cell in row]

        # Compare with the product code
        if prod[1] == prev_prod[1]:
            prod_in_same_category.append(prod)
        else:
            # Sort the product list by trade_time
            prod_in_same_category = sort_by_trade_time(prod_in_same_category)
            generate_by_hour(prod_in_same_category, output_folder)
            generate_by_day(prod_in_same_category, output_folder)
            prod_in_same_category = [prod]
            prev_prod = prod

    # handle the last prod list
    prod_in_same_category = sort_by_trade_time(prod_in_same_category)
    generate_by_hour(prod_in_same_category, output_folder)
    generate_by_day(prod_in_same_category, output_folder)


if __name__ == '__main__':
    main(sys.argv[1:])
